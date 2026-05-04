from collections import Counter
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.accounting import AuditLog
from app.models.documents import Document
from app.models.partners import Partner
from app.models.products import Product
from app.models.stock import StockBalance, StockMovement, Warehouse
from app.schemas.imports import ImportIssue, ImportSummary


TEMPLATES: dict[str, list[str]] = {
    "products": ["sku", "name", "base_price", "description"],
    "partners": ["name", "partner_type", "code", "phone"],
    "warehouses": ["name", "code", "address"],
    "opening-stock": ["product_sku", "product_name", "warehouse_name", "quantity"],
    "opening-partner-balances": ["partner_name", "balance"],
}

REQUIRED: dict[str, list[str]] = {
    "products": ["name"],
    "partners": ["name", "partner_type"],
    "warehouses": ["name"],
    "opening-stock": ["warehouse_name", "quantity"],
    "opening-partner-balances": ["partner_name", "balance"],
}


def build_template(import_type: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = import_type
    for index, column in enumerate(TEMPLATES[import_type], start=1):
        sheet.cell(row=1, column=index, value=column)
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(16, len(column) + 4)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _parse_file(content: bytes, filename: str) -> tuple[list[dict[str, str]], list[ImportIssue]]:
    if filename.lower().endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], [ImportIssue(row=1, field=None, message="File is empty")]
        headers = [str(value or "").strip() for value in rows[0]]
        parsed = []
        for values in rows[1:]:
            parsed.append({header: "" if value is None else str(value).strip() for header, value in zip(headers, values)})
        return parsed, []

    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [{key: (value or "").strip() for key, value in row.items()} for row in reader], []


def _decimal(value: str, row: int, field: str, errors: list[ImportIssue], *, allow_negative: bool = False) -> Decimal | None:
    try:
        number = Decimal(value)
    except (InvalidOperation, TypeError):
        errors.append(ImportIssue(row=row, field=field, message="Invalid number"))
        return None
    if not allow_negative and number < 0:
        errors.append(ImportIssue(row=row, field=field, message="Number must be zero or greater"))
        return None
    return number


def _validate_common(import_type: str, rows: list[dict[str, str]], errors: list[ImportIssue]) -> None:
    for row_index, row in enumerate(rows, start=2):
        for field in REQUIRED[import_type]:
            if not row.get(field):
                errors.append(ImportIssue(row=row_index, field=field, message="Required field is missing"))


def _duplicate_warnings(rows: list[dict[str, str]], key: str) -> list[ImportIssue]:
    counter = Counter(row.get(key, "").strip().lower() for row in rows if row.get(key))
    return [
        ImportIssue(row=index, field=key, message=f"Duplicate {key} in file")
        for index, row in enumerate(rows, start=2)
        if row.get(key) and counter[row[key].strip().lower()] > 1
    ]


def validate_import(db: Session, import_type: str, content: bytes, filename: str) -> tuple[ImportSummary, list[dict[str, str]]]:
    rows, parse_errors = _parse_file(content, filename)
    errors = list(parse_errors)
    warnings: list[ImportIssue] = []
    _validate_common(import_type, rows, errors)

    if import_type == "products":
        warnings.extend(_duplicate_warnings(rows, "sku"))
        warnings.extend(_duplicate_warnings(rows, "name"))
        for index, row in enumerate(rows, start=2):
            if row.get("base_price"):
                _decimal(row["base_price"], index, "base_price", errors)

    if import_type in {"partners", "warehouses"}:
        warnings.extend(_duplicate_warnings(rows, "name"))

    if import_type == "partners":
        for index, row in enumerate(rows, start=2):
            if row.get("partner_type") and row["partner_type"] not in {Partner.TYPE_CUSTOMER, Partner.TYPE_SUPPLIER, Partner.TYPE_BOTH}:
                errors.append(ImportIssue(row=index, field="partner_type", message="Invalid partner_type. Use customer, supplier, or both"))

    if import_type == "opening-stock":
        for index, row in enumerate(rows, start=2):
            product = _find_product(db, row.get("product_sku"), row.get("product_name"))
            warehouse = _find_warehouse(db, row.get("warehouse_name"))
            if product is None:
                errors.append(ImportIssue(row=index, field="product_sku", message="Product not found by sku/name"))
            if warehouse is None:
                errors.append(ImportIssue(row=index, field="warehouse_name", message="Warehouse not found"))
            if row.get("quantity"):
                _decimal(row["quantity"], index, "quantity", errors)

    if import_type == "opening-partner-balances":
        warnings.extend(_duplicate_warnings(rows, "partner_name"))
        for index, row in enumerate(rows, start=2):
            if _find_partner(db, row.get("partner_name")) is None:
                errors.append(ImportIssue(row=index, field="partner_name", message="Partner not found"))
            if row.get("balance"):
                _decimal(row["balance"], index, "balance", errors, allow_negative=True)

    invalid_rows = {issue.row for issue in errors}
    summary = ImportSummary(
        rows_total=len(rows),
        rows_valid=max(0, len(rows) - len(invalid_rows)),
        rows_invalid=len(invalid_rows),
        errors=errors,
        warnings=warnings,
    )
    return summary, rows


def apply_import(db: Session, import_type: str, content: bytes, filename: str) -> ImportSummary:
    summary, rows = validate_import(db, import_type, content, filename)
    if summary.errors:
        return summary
    try:
        created = 0
        skipped = 0
        if import_type == "products":
            created, skipped = _apply_products(db, rows)
        elif import_type == "partners":
            created, skipped = _apply_partners(db, rows)
        elif import_type == "warehouses":
            created, skipped = _apply_warehouses(db, rows)
        elif import_type == "opening-stock":
            created, skipped = _apply_opening_stock(db, rows)
        elif import_type == "opening-partner-balances":
            created, skipped = _apply_opening_partner_balances(db, rows)
        db.add(AuditLog(entity_type="import", entity_id=import_type, action="apply", details=f"created={created};skipped={skipped}"))
        db.commit()
        summary.applied = True
        summary.created = created
        summary.skipped = skipped
        return summary
    except Exception:
        db.rollback()
        raise


def _find_product(db: Session, sku: str | None, name: str | None) -> Product | None:
    if sku:
        product = db.scalar(select(Product).where(Product.sku == sku))
        if product:
            return product
    if name:
        return db.scalar(select(Product).where(Product.name == name))
    return None


def _find_partner(db: Session, name: str | None) -> Partner | None:
    return db.scalar(select(Partner).where(Partner.name == name)) if name else None


def _find_warehouse(db: Session, name: str | None) -> Warehouse | None:
    return db.scalar(select(Warehouse).where(Warehouse.name == name)) if name else None


def _apply_products(db: Session, rows: list[dict[str, str]]) -> tuple[int, int]:
    created = skipped = 0
    for row in rows:
        if _find_product(db, row.get("sku"), row.get("name")):
            skipped += 1
            continue
        db.add(Product(sku=row.get("sku") or None, name=row["name"], base_price=Decimal(row["base_price"] or "0"), description=row.get("description") or None, is_active=True))
        created += 1
    return created, skipped


def _apply_partners(db: Session, rows: list[dict[str, str]]) -> tuple[int, int]:
    created = skipped = 0
    for row in rows:
        if _find_partner(db, row.get("name")):
            skipped += 1
            continue
        db.add(
            Partner(
                name=row["name"],
                partner_type=row["partner_type"],
                code=row.get("code") or None,
                phone=row.get("phone") or None,
                is_active=True,
            )
        )
        created += 1
    return created, skipped


def _apply_warehouses(db: Session, rows: list[dict[str, str]]) -> tuple[int, int]:
    created = skipped = 0
    for row in rows:
        if _find_warehouse(db, row.get("name")):
            skipped += 1
            continue
        db.add(Warehouse(name=row["name"], code=row.get("code") or None, address=row.get("address") or None))
        created += 1
    return created, skipped


def _balance(db: Session, product_id: int, warehouse_id: int) -> StockBalance:
    balance = db.scalar(select(StockBalance).where(StockBalance.product_id == product_id, StockBalance.warehouse_id == warehouse_id))
    if balance is None:
        balance = StockBalance(product_id=product_id, warehouse_id=warehouse_id, quantity=Decimal("0"))
        db.add(balance)
        db.flush()
    return balance


def _apply_opening_stock(db: Session, rows: list[dict[str, str]]) -> tuple[int, int]:
    created = 0
    for row in rows:
        product = _find_product(db, row.get("product_sku"), row.get("product_name"))
        warehouse = _find_warehouse(db, row.get("warehouse_name"))
        assert product is not None and warehouse is not None
        quantity = Decimal(row["quantity"])
        balance = _balance(db, product.id, warehouse.id)
        delta = quantity - balance.quantity
        balance.quantity = quantity
        # TODO LEGACY_RULE_REQUIRED: confirm opening-stock import should be represented as adjustment document or direct movement.
        db.add(StockMovement(product_id=product.id, warehouse_id=warehouse.id, quantity_delta=delta, reason="opening:import"))
        created += 1
    return created, 0


def _apply_opening_partner_balances(db: Session, rows: list[dict[str, str]]) -> tuple[int, int]:
    created = 0
    for row in rows:
        partner = _find_partner(db, row.get("partner_name"))
        assert partner is not None
        balance = Decimal(row["balance"])
        if balance == 0:
            continue
        # TODO LEGACY_RULE_REQUIRED: replace this special posted document with confirmed legacy opening debt representation.
        document_type = Document.TYPE_OUTGOING if balance > 0 else Document.TYPE_INCOMING
        db.add(
            Document(
                document_type=document_type,
                number=f"OPEN-{partner.id}",
                document_date=date.today(),
                status=Document.STATUS_POSTED,
                partner_id=partner.id,
                total_amount=abs(balance),
                note="opening-partner-balance import",
            )
        )
        created += 1
    return created, 0

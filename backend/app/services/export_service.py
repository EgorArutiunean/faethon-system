from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel


@dataclass(frozen=True)
class ExportFile:
    content: bytes
    media_type: str
    filename: str


@dataclass(frozen=True)
class Column:
    key: str
    header: str


REPORT_COLUMNS: dict[str, list[Column]] = {
    "stock-balances": [
        Column("warehouse_name", "Warehouse"),
        Column("product_name", "Product"),
        Column("quantity", "Quantity"),
    ],
    "stock-movements": [
        Column("created_at", "Date"),
        Column("product_name", "Product"),
        Column("warehouse_name", "Warehouse"),
        Column("document_number", "Document"),
        Column("document_type", "Document type"),
        Column("status", "Status"),
        Column("movement_type", "Movement"),
        Column("quantity_delta", "Quantity"),
    ],
    "partner-debts": [
        Column("partner_name", "Partner"),
        Column("partner_type", "Partner type"),
        Column("balance", "Balance"),
    ],
    "cash-book": [
        Column("operation_date", "Date"),
        Column("operation_type", "Type"),
        Column("status", "Status"),
        Column("partner_name", "Partner"),
        Column("payment_id", "Payment ID"),
        Column("document_id", "Document ID"),
        Column("amount", "Amount"),
        Column("cash_balance", "Cash balance"),
    ],
    "documents-register": [
        Column("document_date", "Date"),
        Column("document_number", "Number"),
        Column("document_type", "Type"),
        Column("status", "Status"),
        Column("partner_name", "Partner"),
        Column("warehouse_name", "Warehouse"),
        Column("total_amount", "Total amount"),
    ],
}

REPORT_TITLES = {
    "stock-balances": "Stock balances",
    "stock-movements": "Stock movements",
    "partner-debts": "Partner debts",
    "cash-book": "Cash book",
    "documents-register": "Documents register",
}

REPORT_TOTALS = {
    "stock-balances": [("total_quantity", "Total quantity")],
    "stock-movements": [("total_quantity", "Total quantity")],
    "partner-debts": [("total_partner_debt", "Total partner debt")],
    "cash-book": [
        ("cash_in_total", "Cash in total"),
        ("cash_out_total", "Cash out total"),
        ("cash_balance", "Cash balance"),
    ],
    "documents-register": [("total_amount", "Total amount")],
}


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _row_value(row: BaseModel, key: str) -> object:
    return getattr(row, key, "")


def _active_filters(filters: dict[str, object]) -> list[tuple[str, str]]:
    return [(key, _stringify(value)) for key, value in filters.items() if value not in (None, "", False)]


def _totals(report_key: str, report: BaseModel) -> list[tuple[str, str]]:
    return [(label, _stringify(getattr(report, key))) for key, label in REPORT_TOTALS[report_key]]


def _build_csv(report_key: str, report: BaseModel) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    columns = REPORT_COLUMNS[report_key]
    writer.writerow([column.header for column in columns])
    for row in report.rows:
        writer.writerow([_stringify(_row_value(row, column.key)) for column in columns])
    totals = _totals(report_key, report)
    if totals:
        writer.writerow([])
        for label, value in totals:
            writer.writerow([label, value])
    return output.getvalue().encode("utf-8-sig")


def _build_xlsx(report_key: str, report: BaseModel, filters: dict[str, object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    columns = REPORT_COLUMNS[report_key]

    sheet.cell(row=1, column=1, value=REPORT_TITLES[report_key]).font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value="Generated at")
    sheet.cell(row=2, column=2, value=datetime.now().isoformat(sep=" ", timespec="seconds"))

    row_index = 4
    active_filters = _active_filters(filters)
    if active_filters:
        sheet.cell(row=row_index, column=1, value="Filters").font = Font(bold=True)
        row_index += 1
        for key, value in active_filters:
            sheet.cell(row=row_index, column=1, value=key)
            sheet.cell(row=row_index, column=2, value=value)
            row_index += 1
        row_index += 1

    for column_index, column in enumerate(columns, start=1):
        sheet.cell(row=row_index, column=column_index, value=column.header).font = Font(bold=True)
    row_index += 1

    for row in report.rows:
        for column_index, column in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=_stringify(_row_value(row, column.key)))
        row_index += 1

    row_index += 1
    for label, value in _totals(report_key, report):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)
        row_index += 1

    for column_cells in sheet.columns:
        max_length = max(len(_stringify(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_report(report_key: str, report: BaseModel, export_format: str, filters: dict[str, object]) -> ExportFile:
    if export_format == "csv":
        return ExportFile(
            content=_build_csv(report_key, report),
            media_type="text/csv; charset=utf-8",
            filename=f"{report_key}.csv",
        )
    if export_format == "xlsx":
        return ExportFile(
            content=_build_xlsx(report_key, report, filters),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{report_key}.xlsx",
        )
    raise ValueError("Unsupported export format")

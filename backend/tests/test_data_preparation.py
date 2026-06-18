from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
import csv
import sys

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).resolve().parents[2] / "migration" / "data_preparation" / "prepare_legacy_price_list.py"
SPEC = spec_from_file_location("prepare_legacy_price_list", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
prepare_legacy_price_list = module_from_spec(SPEC)
sys.modules[SPEC.name] = prepare_legacy_price_list
SPEC.loader.exec_module(prepare_legacy_price_list)


def test_prepare_legacy_price_list_writes_standard_import_templates(tmp_path: Path) -> None:
    input_path = tmp_path / "price.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Price list"])
    sheet.append(
        [
            "\u0421\u043a\u043b\u0430\u0434",
            "\u041a\u043e\u0434",
            "\u0422\u043e\u0432\u0430\u0440",
            "\u0415\u0434.",
            "\u041a\u043e\u043b-\u0432\u043e",
            "\u0426\u0435\u043d\u0430 \u043e\u0441\u0442.",
        ]
    )
    legacy_name = "\u041a\u0440\u0443\u043f\u0430 \u041a\u0443\u0442\u044c\u044f 0,9\u043a\u0433*17\u0448\u0442"
    sheet.append(["MAIN", "93197", legacy_name, "\u0448\u0442", 17, "136.00"])
    workbook.save(input_path)

    category_map_path = tmp_path / "category-map.csv"
    with category_map_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["legacy_name", "category"])
        writer.writeheader()
        writer.writerow({"legacy_name": legacy_name, "category": "\u0411\u0430\u043a\u0430\u043b\u0435\u044f"})

    rows = prepare_legacy_price_list.prepare_rows(input_path, category_map_path)
    products_path, stock_path = prepare_legacy_price_list.write_standard_imports(rows, tmp_path / "out", "MAIN")

    products_sheet = load_workbook(products_path, read_only=True).active
    product_rows = list(products_sheet.iter_rows(values_only=True))
    assert product_rows[0] == ("sku", "name", "category", "base_price", "description", "legacy_name")
    assert product_rows[1] == ("93197", legacy_name, "\u0411\u0430\u043a\u0430\u043b\u0435\u044f", "136.00", None, legacy_name)

    stock_sheet = load_workbook(stock_path, read_only=True).active
    stock_rows = list(stock_sheet.iter_rows(values_only=True))
    assert stock_rows[0] == ("product_sku", "product_name", "warehouse_name", "quantity")
    assert stock_rows[1] == ("93197", legacy_name, "MAIN", "17")

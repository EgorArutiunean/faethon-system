from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


PRODUCT_COLUMNS = ["sku", "name", "category", "base_price", "description", "legacy_name"]
OPENING_STOCK_COLUMNS = ["product_sku", "product_name", "warehouse_name", "quantity"]

OLD_CODE = "\u041a\u043e\u0434"
OLD_NAME = "\u0422\u043e\u0432\u0430\u0440"
OLD_QUANTITY = "\u041a\u043e\u043b-\u0432\u043e"
OLD_REST_PRICE = "\u0426\u0435\u043d\u0430 \u043e\u0441\u0442."


@dataclass(frozen=True)
class PreparedRow:
    sku: str
    name: str
    category: str
    base_price: str
    legacy_name: str
    quantity: str


def _string(value: object) -> str:
    return "" if value is None else str(value).strip()


def _decimal_string(value: object) -> str:
    text = _string(value).replace(",", ".")
    if not text:
        return ""
    try:
        return str(Decimal(text))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc


def _header_score(values: Iterable[object]) -> int:
    headers = {_string(value) for value in values}
    return sum(1 for name in [OLD_CODE, OLD_NAME, OLD_QUANTITY, OLD_REST_PRICE] if name in headers)


def _detect_header_row(rows: list[tuple[object, ...]]) -> int:
    best_index = 0
    best_score = 0
    for index, row in enumerate(rows[:10]):
        score = _header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    if best_score < 2:
        raise ValueError("Could not find old price-list header row")
    return best_index


def _load_category_map(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"legacy_name", "category"}
        if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
            raise ValueError("Category map must contain legacy_name and category columns")
        return {row["legacy_name"].strip(): row["category"].strip() for row in reader if row.get("legacy_name") and row.get("category")}


def prepare_rows(input_path: Path, category_map_path: Path | None = None) -> list[PreparedRow]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Input workbook is empty")

    category_map = _load_category_map(category_map_path)
    header_index = _detect_header_row(rows)
    headers = [_string(value) for value in rows[header_index]]
    index_by_header = {header: index for index, header in enumerate(headers) if header}

    missing = [header for header in [OLD_CODE, OLD_NAME] if header not in index_by_header]
    if missing:
        raise ValueError(f"Required old price-list columns are missing: {', '.join(missing)}")

    prepared: list[PreparedRow] = []
    for values in rows[header_index + 1 :]:
        sku = _string(values[index_by_header[OLD_CODE]] if index_by_header[OLD_CODE] < len(values) else "")
        legacy_name = _string(values[index_by_header[OLD_NAME]] if index_by_header[OLD_NAME] < len(values) else "")
        if not sku and not legacy_name:
            continue
        quantity = ""
        if OLD_QUANTITY in index_by_header and index_by_header[OLD_QUANTITY] < len(values):
            quantity = _decimal_string(values[index_by_header[OLD_QUANTITY]])
        base_price = ""
        if OLD_REST_PRICE in index_by_header and index_by_header[OLD_REST_PRICE] < len(values):
            base_price = _decimal_string(values[index_by_header[OLD_REST_PRICE]])
        prepared.append(
            PreparedRow(
                sku=sku,
                name=legacy_name,
                category=category_map.get(legacy_name, ""),
                base_price=base_price,
                legacy_name=legacy_name,
                quantity=quantity,
            )
        )
    return prepared


def _write_workbook(path: Path, columns: list[str], data_rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = path.stem
    sheet.append(columns)
    for row in data_rows:
        sheet.append(row)
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(16, len(column) + 4)
    workbook.save(path)


def write_standard_imports(rows: list[PreparedRow], output_dir: Path, warehouse_name: str) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    products_path = output_dir / "products.xlsx"
    stock_path = output_dir / "opening-stock.xlsx"

    _write_workbook(
        products_path,
        PRODUCT_COLUMNS,
        [[row.sku, row.name, row.category, row.base_price, "", row.legacy_name] for row in rows],
    )
    _write_workbook(
        stock_path,
        OPENING_STOCK_COLUMNS,
        [[row.sku, row.name, warehouse_name, row.quantity] for row in rows if row.quantity],
    )
    return products_path, stock_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare old price-list data into Buy Modern Import Lite templates.")
    parser.add_argument("--input", required=True, type=Path, help="Path to the old price-list XLSX file.")
    parser.add_argument("--output-dir", required=True, type=Path, help="Directory for generated Buy Modern templates.")
    parser.add_argument("--warehouse-name", required=True, help="Warehouse name to put into opening-stock.xlsx.")
    parser.add_argument("--category-map", type=Path, help="Optional CSV with legacy_name,category columns.")
    args = parser.parse_args()

    rows = prepare_rows(args.input, args.category_map)
    products_path, stock_path = write_standard_imports(rows, args.output_dir, args.warehouse_name)
    print(f"Prepared {len(rows)} products")
    print(f"Products: {products_path}")
    print(f"Opening stock: {stock_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
